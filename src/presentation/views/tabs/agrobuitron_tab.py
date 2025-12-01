"""
Agrobuitron Tab - Invoice processing interface for Agrobuitron company
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QRadioButton, QButtonGroup, QFileDialog,
    QComboBox, QProgressBar, QMessageBox, QFrame, QScrollArea, QSizePolicy
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from typing import List, Optional
from openpyxl import load_workbook
import subprocess
import platform
from pathlib import Path


class ProcessingThread(QThread):
    """Thread for processing invoices without blocking UI"""

    progress_update = pyqtSignal(int, int)  # current, total
    finished = pyqtSignal(bool, str, int)  # success, message, records

    def __init__(self, controller, zip_files, company, output_format, excel_file=None, excel_sheet=None):
        super().__init__()
        self.controller = controller
        self.zip_files = zip_files
        self.company = company
        self.output_format = output_format
        self.excel_file = excel_file
        self.excel_sheet = excel_sheet

    def run(self):
        """Run processing in background thread"""
        success, message, records = self.controller.process_invoices(
            zip_files=self.zip_files,
            company=self.company,
            output_format=self.output_format,
            excel_file=self.excel_file,
            excel_sheet=self.excel_sheet,
            progress_callback=self.progress_update.emit
        )
        self.finished.emit(success, message, records)


class AgrobuitronTab(QWidget):
    """Tab for Agrobuitron invoice processing"""

    def __init__(self, main_controller):
        super().__init__()
        self.main_controller = main_controller
        self.zip_files: List[str] = []
        self.excel_file: Optional[str] = None
        self.processing_thread: Optional[ProcessingThread] = None
        self.card_style = (
            "QFrame { background-color: white; border-radius: 8px; "
            "padding: 15px; border: 1px solid #e0e0e0; }"
        )
        self.init_ui()

    def init_ui(self):
        """Initialize the user interface"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet("QScrollArea { border: 0; }")
        main_layout.addWidget(scroll_area)

        content_widget = QWidget()
        content_widget.setStyleSheet("background-color: #f5f7fa;")
        scroll_area.setWidget(content_widget)

        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # Company title
        title = QLabel("AGROBUITRON")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title.setStyleSheet("color: #2E8B57;")
        layout.addWidget(title)

        # ZIP files section
        layout.addWidget(self._create_zip_section())

        # Output format section
        layout.addWidget(self._create_output_section())

        # Excel options section (initially hidden)
        self.excel_frame = self._create_excel_section()
        self.excel_frame.hide()
        layout.addWidget(self.excel_frame)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
            }
        """)
        layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        self.status_label.setStyleSheet("color: #7f8c8d; font-size: 10pt;")
        layout.addWidget(self.status_label)

        # Process button
        process_btn = QPushButton("PROCESAR FACTURAS")
        process_btn.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        process_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 15px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        process_btn.clicked.connect(self.process_invoices)
        process_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        layout.addWidget(process_btn)

        layout.addStretch()
        self.setLayout(main_layout)

    def _create_zip_section(self) -> QFrame:
        """Create ZIP files selection section"""
        frame = QFrame()
        frame.setStyleSheet(self.card_style)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        label = QLabel("Archivos ZIP:")
        label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(label)

        self.zip_list = QListWidget()
        self.zip_list.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.zip_list.setMinimumHeight(180)
        layout.addWidget(self.zip_list)

        btn_layout = QHBoxLayout()

        add_btn = QPushButton("Agregar ZIP")
        add_btn.clicked.connect(self.add_zip_files)
        add_btn.setStyleSheet(self._get_button_style("#3498db"))
        btn_layout.addWidget(add_btn)

        remove_btn = QPushButton("Quitar Seleccionado")
        remove_btn.clicked.connect(self.remove_zip_file)
        remove_btn.setStyleSheet(self._get_button_style("#e74c3c"))
        btn_layout.addWidget(remove_btn)

        clear_btn = QPushButton("Limpiar Todo")
        clear_btn.clicked.connect(self.clear_zip_files)
        clear_btn.setStyleSheet(self._get_button_style("#95a5a6"))
        btn_layout.addWidget(clear_btn)

        layout.addLayout(btn_layout)
        frame.setLayout(layout)
        return frame

    def _create_output_section(self) -> QFrame:
        """Create output format selection section"""
        frame = QFrame()
        frame.setStyleSheet(self.card_style)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        label = QLabel("Formato de Salida:")
        label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(label)

        self.output_group = QButtonGroup()

        self.csv_radio = QRadioButton("Generar nuevo archivo CSV")
        self.csv_radio.setChecked(True)
        self.csv_radio.toggled.connect(self._on_output_format_changed)
        self.output_group.addButton(self.csv_radio)
        layout.addWidget(self.csv_radio)

        self.excel_radio = QRadioButton("Actualizar archivo Excel existente")
        self.excel_radio.toggled.connect(self._on_output_format_changed)
        self.output_group.addButton(self.excel_radio)
        layout.addWidget(self.excel_radio)

        frame.setLayout(layout)
        return frame

    def _create_excel_section(self) -> QFrame:
        """Create Excel file selection section"""
        frame = QFrame()
        frame.setStyleSheet(self.card_style)

        layout = QVBoxLayout()
        layout.setSpacing(10)

        # Excel file selection label
        file_label = QLabel("Archivo Excel:")
        file_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(file_label)

        # Path and button in horizontal layout
        file_row = QHBoxLayout()
        self.excel_path_label = QLabel("Ninguno seleccionado")
        self.excel_path_label.setStyleSheet("color: #7f8c8d;")
        self.excel_path_label.setWordWrap(True)
        self.excel_path_label.setMinimumWidth(300)
        file_row.addWidget(self.excel_path_label, 1)

        select_excel_btn = QPushButton("Seleccionar...")
        select_excel_btn.setMinimumWidth(120)
        select_excel_btn.clicked.connect(self.select_excel_file)
        select_excel_btn.setStyleSheet(self._get_button_style("#3498db"))
        file_row.addWidget(select_excel_btn)

        layout.addLayout(file_row)

        # Sheet selection
        sheet_label = QLabel("Hoja:")
        sheet_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(sheet_label)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumHeight(35)
        layout.addWidget(self.sheet_combo)

        frame.setLayout(layout)
        return frame

    def add_zip_files(self):
        """Add ZIP files to the list"""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Seleccionar Archivos ZIP",
            "",
            "ZIP Files (*.zip)"
        )

        if files:
            for file in files:
                if file not in self.zip_files:
                    self.zip_files.append(file)
                    self.zip_list.addItem(file)

    def remove_zip_file(self):
        """Remove selected ZIP file"""
        current_row = self.zip_list.currentRow()
        if current_row >= 0:
            self.zip_files.pop(current_row)
            self.zip_list.takeItem(current_row)

    def clear_zip_files(self):
        """Clear all ZIP files"""
        self.zip_files.clear()
        self.zip_list.clear()

    def select_excel_file(self):
        """Select Excel file"""
        file, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Archivo Excel",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file:
            self.excel_file = file
            self.excel_path_label.setText(file)
            self._load_excel_sheets()

    def _load_excel_sheets(self):
        """Load sheets from Excel file"""
        if not self.excel_file:
            return

        try:
            wb = load_workbook(self.excel_file, read_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(wb.sheetnames)
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error cargando hojas de Excel: {str(e)}")

    def _on_output_format_changed(self):
        """Handle output format change"""
        if self.excel_radio.isChecked():
            self.excel_frame.show()
        else:
            self.excel_frame.hide()

    def process_invoices(self):
        """Process invoice ZIP files"""
        if not self.zip_files:
            QMessageBox.warning(self, "Error", "Por favor seleccione al menos un archivo ZIP")
            return

        output_format = 'csv' if self.csv_radio.isChecked() else 'excel'

        if output_format == 'excel' and not self.excel_file:
            QMessageBox.warning(self, "Error", "Por favor seleccione un archivo Excel")
            return

        # Disable UI during processing
        self.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("Procesando facturas...")

        # Start processing in background thread
        self.processing_thread = ProcessingThread(
            self.main_controller,
            self.zip_files,
            "AGROBUITRON",
            output_format,
            self.excel_file if output_format == 'excel' else None,
            self.sheet_combo.currentText() if output_format == 'excel' else None
        )

        self.processing_thread.progress_update.connect(self._on_progress_update)
        self.processing_thread.finished.connect(self._on_processing_finished)
        self.processing_thread.start()

    def _on_progress_update(self, current: int, total: int):
        """Handle progress update"""
        if total > 0:
            progress = int((current / total) * 100)
            self.progress_bar.setValue(progress)

    def _on_processing_finished(self, success: bool, message: str, records: int):
        """Handle processing completion"""
        self.setEnabled(True)
        self.progress_bar.setValue(100 if success else 0)
        self.status_label.setText(f"Proceso completado. Registros procesados: {records}" if success else "Error en el procesamiento")

        if success:
            self._show_success_dialog(message)
        else:
            QMessageBox.critical(self, "Error", message)

    def _show_success_dialog(self, message: str):
        """Show success dialog with option to open folder"""
        # Extract file path from message
        file_path = self._extract_file_path_from_message(message)

        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("Éxito")
        msg_box.setText(message)

        # Add custom buttons
        msg_box.addButton("OK", QMessageBox.ButtonRole.AcceptRole)

        if file_path:
            open_folder_btn = msg_box.addButton("Abrir carpeta", QMessageBox.ButtonRole.ActionRole)
            msg_box.exec()

            # Check which button was clicked
            if msg_box.clickedButton() == open_folder_btn:
                self._open_file_location(file_path)
        else:
            msg_box.exec()

    def _extract_file_path_from_message(self, message: str) -> Optional[str]:
        """Extract file path from success message"""
        # Messages typically contain the path after a colon or newline
        lines = message.split('\n')
        for line in lines:
            line = line.strip()
            if line and (line.startswith('/') or line.startswith('data/')):
                # Check if it's a valid file path
                path = Path(line)
                if path.exists():
                    return str(path.absolute())
        return None

    def _open_file_location(self, file_path: str):
        """Open the folder containing the file in the system file explorer"""
        try:
            path = Path(file_path)
            folder_path = path.parent if path.is_file() else path

            system = platform.system()
            if system == "Windows":
                subprocess.run(['explorer', '/select,', str(path.absolute())])
            elif system == "Darwin":  # macOS
                subprocess.run(['open', '-R', str(path.absolute())])
            else:  # Linux and others
                subprocess.run(['xdg-open', str(folder_path.absolute())])
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo abrir la carpeta: {str(e)}")

    def _get_button_style(self, color: str) -> str:
        """Get button stylesheet"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self._darken_color(color)};
            }}
        """

    def _darken_color(self, color: str, factor: float = 0.9) -> str:
        """Darken a hex color"""
        color = color.lstrip('#')
        rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        darkened = tuple(int(c * factor) for c in rgb)
        return f"#{darkened[0]:02x}{darkened[1]:02x}{darkened[2]:02x}"
