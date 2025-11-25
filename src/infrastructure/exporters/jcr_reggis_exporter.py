
"""
Reggis XLSX Exporter for Juan Camilo Rosas
Exports invoices to the specific Reggis format in Excel
"""
from pathlib import Path
from datetime import datetime
from typing import List
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from ...domain.entities.invoice import Invoice


class JCRReggisExporter:
    """
    Exports Juan Camilo Rosas invoices to Reggis XLSX format

    Reggis Format Columns:
    N? Factura, Nombre Producto, Codigo Subyacente, Unidad Medida en Kg,Un,Lt,
    Cantidad (5 decimales - separador coma), Precio Unitario (5 decimales - separador coma),
    Fecha Factura A?o-Mes-Dia, Fecha Pago A?o-Mes-Dia, Nit Comprador (Existente),
    Nombre Comprador, Nit Vendedor (Existente), Nombre Vendedor, Principal V,C,
    Municipio (Nombre Exacto de la Ciudad), Iva (N?%), Descripci?n, Activa,
    Factura Activa, Bodega, Incentivo, Cantidad Original (5 decimales - separador coma),
    Moneda (1,2,3), Valor Total
    """

    # Reggis column order
    REGGIS_COLUMNS = [
        'N? Factura',
        'Nombre Producto',
        'Codigo Subyacente',
        'Unidad Medida en Kg,Un,Lt',
        'Cantidad (5 decimales - separador coma)',
        'Precio Unitario (5 decimales - separador coma)',
        'Fecha Factura A?o-Mes-Dia',
        'Fecha Pago A?o-Mes-Dia',
        'Nit Comprador (Existente)',
        'Nombre Comprador',
        'Nit Vendedor (Existente)',
        'Nombre Vendedor',
        'Principal V,C',
        'Municipio (Nombre Exacto de la Ciudad)',
        'Iva (N?%)',
        'Descripci?n',
        'Activa',
        'Factura Activa',
        'Bodega',
        'Incentivo',
        'Cantidad Original (5 decimales - separador coma)',
        'Moneda (1,2,3)',
        'Valor Total'
    ]

    def __init__(self, municipality: str = '', iva_percentage: str = '0'):
        """
        Initialize exporter

        Args:
            municipality: Municipality name for all invoices
            iva_percentage: Default IVA percentage
        """
        self.municipality = municipality
        self.iva_percentage = iva_percentage

    def export_to_reggis_csv(
        self,
        invoices: List[Invoice],
        original_quantities: dict = None,
        company: str = "JUAN CAMILO ROSAS"
    ) -> str:
        """
        Export invoices to Reggis XLSX format

        Args:
            invoices: List of Invoice entities
            original_quantities: Dictionary mapping (invoice_number, product_name) to original quantity
            company: Company name used for output folder/name

        Returns:
            Path to the generated XLSX file
        """
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        date_folder = now.strftime("%d-%m-%Y")
        safe_company = company if company else "JUAN CAMILO ROSAS"
        prefix = safe_company.replace(" ", "_")
        filename = f"{prefix}_Reggis_Facturas_{timestamp}.xlsx"
        output_dir = Path("data") / safe_company / date_folder / "archivos"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / filename

        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas Reggis"

        # Write header row with bold font
        for col_idx, column_name in enumerate(self.REGGIS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        row_idx = 2
        for invoice in invoices:
            for product in invoice.products:
                # Get original quantity if available
                key = (invoice.invoice_number, product.name)
                if original_quantities and key in original_quantities:
                    original_qty = original_quantities[key]
                else:
                    # If not available, use the converted quantity
                    original_qty = product.quantity

                # Prepare row data
                row_data = [
                    invoice.invoice_number,
                    product.name,
                    'SPN-1',
                    product.unit_of_measure,
                    self._format_decimal(product.quantity),
                    self._format_decimal(product.unit_price),
                    invoice.get_issue_date_formatted(),
                    invoice.get_due_date_formatted(),
                    invoice.buyer_nit,
                    invoice.buyer_name,
                    invoice.seller_nit,  # Always '1003516945'
                    invoice.seller_name,  # Always 'JUAN CAMILO ROSAS'
                    'V',  # Always 'V'
                    self.municipality if self.municipality else invoice.seller_municipality,
                    f"{product.get_formatted_iva()}%",  # IVA del producto
                    '',  # Descripción - Always empty
                    '1',  # Activa - Always 1
                    '1',  # Factura Activa - Always 1
                    '',  # Bodega
                    '',  # Incentivo
                    self._format_decimal(original_qty),
                    '1',  # Moneda - Always 1 (COP)
                    self._format_decimal(product.total_price)
                ]

                # Write row
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

                row_idx += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)

        return str(output_path.resolve())

    def _format_decimal(self, value) -> str:
        """
        Format decimal with 5 decimals and comma as decimal separator

        Args:
            value: Decimal or numeric value

        Returns:
            Formatted string with comma as decimal separator
        """
        if isinstance(value, str):
            try:
                value = Decimal(value)
            except:
                return "0,00000"

        # Format with 5 decimals
        formatted = f"{float(value):.5f}"

        # Replace dot with comma
        return formatted.replace('.', ',')
