"""
Excel Exporter - Exports invoices to Excel format
"""
from typing import List, Dict, Optional
from openpyxl import load_workbook

from ...domain.entities.invoice import Invoice


class ExcelExporter:
    """Exports data to Excel format by updating existing workbooks"""

    def export_to_excel(
        self,
        invoices: List[Invoice],
        excel_file: str,
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Export invoices to an existing Excel file

        Args:
            invoices: List of Invoice entities
            excel_file: Path to existing Excel file
            sheet_name: Name of the sheet to update (uses active sheet if None)
        """
        # Load workbook
        wb = load_workbook(excel_file)

        # Get worksheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read headers from first row
        headers = self._read_headers(ws)

        # Get field mapping for flexible column matching
        field_mapping = self._get_field_mapping()

        # Find column positions for each field
        column_positions = {}
        for field, possible_names in field_mapping.items():
            for col_name, col_idx in headers.items():
                if col_name in possible_names:
                    column_positions[field] = col_idx
                    break

        # Find first empty row
        first_empty_row = ws.max_row + 1
        for row_idx in range(2, ws.max_row + 2):
            if ws.cell(row=row_idx, column=1).value is None:
                first_empty_row = row_idx
                break

        # Write invoice data
        current_row = first_empty_row
        for invoice in invoices:
            for product in invoice.products:
                # Prepare row data
                row_data = {
                    'N° Factura': invoice.invoice_number,
                    'Nombre Producto': product.name,
                    'Codigo Subyacente': product.underlying_code,
                    'Unidad Medida': product.unit_of_measure,
                    'Cantidad': product.get_formatted_quantity(),
                    'Precio Unitario': product.get_formatted_unit_price(),
                    'Precio Total': product.get_formatted_total_price(),
                    'Fecha Factura': invoice.get_issue_date_formatted(),
                    'Fecha Pago': invoice.get_due_date_formatted(),
                    'Nit Comprador': invoice.buyer_nit,
                    'Nombre Comprador': invoice.buyer_name,
                    'Nit Vendedor': invoice.seller_nit,
                    'Nombre Vendedor': invoice.seller_name,
                    'Principal V,C': 'V',
                    'Municipio': invoice.seller_municipality,
                    'Iva': f"{product.iva_percentage}%",
                    'Descripción': '',
                    'Activa Factura': 'Sí',
                    'Activa Bodega': 'Sí',
                    'Incentivo': '',
                    'Cantidad Original': product.get_formatted_quantity(),
                    'Moneda': invoice.format_currency_code()
                }

                # Write data to columns
                for field, value in row_data.items():
                    if field in column_positions:
                        col_idx = column_positions[field]
                        ws.cell(row=current_row, column=col_idx, value=value)

                current_row += 1

        # Save workbook
        wb.save(excel_file)

    def _read_headers(self, worksheet) -> Dict[str, int]:
        """
        Read headers from first row of worksheet

        Args:
            worksheet: openpyxl worksheet

        Returns:
            Dictionary mapping header name to column index
        """
        headers = {}
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx

        return headers

    def _get_field_mapping(self) -> Dict[str, List[str]]:
        """
        Get field mapping with possible column names

        Returns:
            Dictionary mapping field name to list of possible column names
        """
        return {
            "N° Factura": ["N° Factura", "Numero Factura", "No. Factura", "Factura"],
            "Nombre Producto": ["Nombre Producto", "Producto", "Descripcion Producto"],
            "Codigo Subyacente": ["Codigo Subyacente", "Código Subyacente", "Codigo", "Code"],
            "Unidad Medida": ["Unidad Medida", "Unidad", "U/M", "UM"],
            "Cantidad": ["Cantidad", "Qty", "Cantidad Original"],
            "Precio Unitario": ["Precio Unitario", "Precio Unit", "P.Unit", "Unit Price"],
            "Precio Total": ["Precio Total", "Total", "Price Total"],
            "Fecha Factura": ["Fecha Factura", "Fecha", "Date", "Issue Date"],
            "Fecha Pago": ["Fecha Pago", "Due Date", "Vencimiento"],
            "Nit Comprador": ["Nit Comprador", "NIT Comprador", "Customer NIT"],
            "Nombre Comprador": ["Nombre Comprador", "Comprador", "Customer"],
            "Nit Vendedor": ["Nit Vendedor", "NIT Vendedor", "Supplier NIT"],
            "Nombre Vendedor": ["Nombre Vendedor", "Vendedor", "Supplier"],
            "Principal V,C": ["Principal V,C", "Principal", "Type"],
            "Municipio": ["Municipio", "Ciudad", "City"],
            "Iva": ["Iva", "IVA", "Tax", "VAT"],
            "Descripción": ["Descripción", "Descripcion", "Description", "Desc"],
            "Activa Factura": ["Activa Factura", "Active Invoice"],
            "Activa Bodega": ["Activa Bodega", "Active Warehouse"],
            "Incentivo": ["Incentivo", "Incentive"],
            "Cantidad Original": ["Cantidad Original", "Original Qty"],
            "Moneda": ["Moneda", "Currency"]
        }
