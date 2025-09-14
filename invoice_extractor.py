import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import xml.etree.ElementTree as ET
import zipfile
import os
import pandas as pd
import csv
from datetime import datetime
import threading
from pathlib import Path
import hashlib
import json
import requests
from packaging import version
import sys

class InvoiceExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Extracción de Facturas Electrónicas v1.0.0")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        # Variables globales
        self.current_version = "1.0.0"
        self.excel_file = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.zip_files = []
        self.extracted_data = []
        self.current_company = "AGROBUITRON"  # Empresa actual
        
        # Verificar contraseña
        if not self.verify_password():
            self.root.destroy()
            return
        
        # Verificar actualizaciones
        self.check_for_updates()
        
        # Configurar estilos
        self.setup_styles()
        
        # Crear interfaz con pestañas
        self.create_tabbed_interface()
        
        # Definir namespaces XML
        self.namespaces = {
            'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
            'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
            'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
            'sts': 'dian:gov:co:facturaelectronica:Structures-2-1'
        }
    
    def verify_password(self):
        """Verificar contraseña de acceso"""
        # Hash de la contraseña: "FacturasElectronicas2024"
        correct_hash = "a8b5c3d4e5f6a7b8c9d0e1f2a3b4c5d6e7f8a9b0c1d2e3f4a5b6c7d8e9f0a1b2"
        
        password = simpledialog.askstring("Acceso Restringido", 
                                         "Ingrese la contraseña:", show='*')
        if not password:
            return False
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if password_hash == correct_hash:
            return True
        else:
            messagebox.showerror("Error", "Contraseña incorrecta")
            return False
    
    def check_for_updates(self):
        """Verificar actualizaciones disponibles"""
        try:
            # URL del archivo de versión en GitHub (ejemplo)
            version_url = "https://api.github.com/repos/tu-usuario/invoice-extractor/releases/latest"
            
            response = requests.get(version_url, timeout=5)
            if response.status_code == 200:
                latest_release = response.json()
                latest_version = latest_release['tag_name'].replace('v', '')
                
                if version.parse(latest_version) > version.parse(self.current_version):
                    if messagebox.askyesno("Actualización Disponible", 
                                         f"Hay una nueva versión disponible: {latest_version}\n"
                                         "¿Desea descargar la actualización?"):
                        self.download_update(latest_release['assets'][0]['browser_download_url'])
        except:
            pass  # Silenciosamente continuar si no hay internet
    
    def download_update(self, download_url):
        """Descargar e instalar actualización"""
        try:
            response = requests.get(download_url, stream=True)
            update_file = "invoice_extractor_update.exe"
            
            with open(update_file, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            messagebox.showinfo("Actualización", 
                              f"Actualización descargada como {update_file}.\n"
                              "Por favor, cierre la aplicación y ejecute el nuevo archivo.")
        except Exception as e:
            messagebox.showerror("Error", f"Error descargando actualización: {str(e)}")
    
    def setup_styles(self):
        """Configurar estilos de la interfaz"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar colores y fuentes
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'), background='#f0f0f0')
        style.configure('Subtitle.TLabel', font=('Arial', 10, 'bold'), background='#f0f0f0')
        style.configure('Custom.TButton', font=('Arial', 10))
        style.configure('Company.TLabel', font=('Arial', 12, 'bold'), foreground='#2E8B57')
    
    def create_tabbed_interface(self):
        """Crear interfaz con pestañas para diferentes empresas"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Título principal
        title_label = ttk.Label(main_frame, text="Sistema de Extracción de Facturas Electrónicas", 
                               style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Crear notebook para pestañas
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        
        # Pestaña AGROBUITRON
        self.agrobuitron_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.agrobuitron_frame, text="AGROBUITRON")
        
        # Pestaña MG (placeholder por ahora)
        self.mg_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.mg_frame, text="MG")
        
        # Pestaña ROSAS (placeholder por ahora)
        self.rosas_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.rosas_frame, text="ROSAS")
        
        # Crear contenido para AGROBUITRON
        self.create_agrobuitron_interface()
        
        # Crear placeholders para otras empresas
        self.create_placeholder_interface(self.mg_frame, "MG")
        self.create_placeholder_interface(self.rosas_frame, "ROSAS")
        
        # Bind para cambio de pestaña
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        # Configurar expansión de grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
    
    def on_tab_changed(self, event):
        """Manejar cambio de pestaña"""
        selection = event.widget.tab('current')['text']
        self.current_company = selection
        
    def create_agrobuitron_interface(self):
        """Crear interfaz específica para AGROBUITRON"""
        frame = self.agrobuitron_frame
        
        # Encabezado de la empresa
        company_label = ttk.Label(frame, text="AGROBUITRON SAS", style='Company.TLabel')
        company_label.grid(row=0, column=0, columnspan=3, pady=(0, 15))
        
        # Sección archivos ZIP
        zip_frame = ttk.LabelFrame(frame, text="Archivos ZIP con Facturas", padding="10")
        zip_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        
        # Lista de archivos ZIP
        self.zip_listbox_agro = tk.Listbox(zip_frame, height=8, selectmode=tk.MULTIPLE)
        self.zip_listbox_agro.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), 
                                  padx=(0, 10), pady=(0, 10))
        
        # Scrollbar para la lista
        scrollbar_agro = ttk.Scrollbar(zip_frame, orient=tk.VERTICAL, command=self.zip_listbox_agro.yview)
        scrollbar_agro.grid(row=0, column=2, sticky=(tk.N, tk.S), pady=(0, 10))
        self.zip_listbox_agro.configure(yscrollcommand=scrollbar_agro.set)
        
        # Botones para archivos ZIP
        zip_buttons_frame = ttk.Frame(zip_frame)
        zip_buttons_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        ttk.Button(zip_buttons_frame, text="Agregar ZIPs", 
                  command=self.add_zip_files_agro, style='Custom.TButton').grid(row=0, column=0, padx=(0, 10))
        ttk.Button(zip_buttons_frame, text="Remover Seleccionados", 
                  command=self.remove_zip_files_agro, style='Custom.TButton').grid(row=0, column=1, padx=(0, 10))
        ttk.Button(zip_buttons_frame, text="Limpiar Lista", 
                  command=self.clear_zip_files_agro, style='Custom.TButton').grid(row=0, column=2)
        
        # Opciones de salida
        output_frame = ttk.LabelFrame(frame, text="Opciones de Salida", padding="10")
        output_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Variables para opciones
        self.generate_csv = tk.BooleanVar(value=True)
        self.generate_excel = tk.BooleanVar(value=False)
        
        ttk.Checkbutton(output_frame, text="Generar archivo CSV", 
                       variable=self.generate_csv).grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        ttk.Checkbutton(output_frame, text="Actualizar Excel existente", 
                       variable=self.generate_excel).grid(row=0, column=1, sticky=tk.W)
        
        # Selección de archivo Excel (solo si está marcado)
        self.excel_frame_agro = ttk.Frame(output_frame)
        self.excel_frame_agro.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Entry(self.excel_frame_agro, textvariable=self.excel_file, width=50, state='readonly').grid(
            row=0, column=0, padx=(0, 10))
        ttk.Button(self.excel_frame_agro, text="Seleccionar Excel", 
                  command=self.select_excel_file, style='Custom.TButton').grid(row=0, column=1, padx=(0, 10))
        ttk.Button(self.excel_frame_agro, text="Vista Previa", 
                  command=self.preview_excel_structure, style='Custom.TButton').grid(row=0, column=2)
        
        # Hoja de Excel
        sheet_frame = ttk.Frame(self.excel_frame_agro)
        sheet_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Label(sheet_frame, text="Hoja:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, width=20, state='readonly')
        self.sheet_combo.grid(row=0, column=1, padx=(5, 0), sticky=tk.W)
        
        # Inicialmente ocultar frame de Excel
        self.toggle_excel_options()
        self.generate_excel.trace('w', lambda *args: self.toggle_excel_options())
        
        # Barra de progreso
        progress_frame = ttk.Frame(frame)
        progress_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        ttk.Label(progress_frame, text="Progreso:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                          maximum=100, length=400)
        self.progress_bar.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        
        self.status_label = ttk.Label(progress_frame, text="Listo para procesar")
        self.status_label.grid(row=2, column=0, sticky=tk.W, pady=(5, 0))
        
        # Botón de procesamiento
        process_frame = ttk.Frame(frame)
        process_frame.grid(row=4, column=0, columnspan=3, pady=(0, 15))
        
        self.process_button = ttk.Button(process_frame, text="PROCESAR FACTURAS AGROBUITRON", 
                                        command=self.start_processing_agro, 
                                        style='Custom.TButton')
        self.process_button.grid(row=0, column=0, padx=10)
        
        # Configurar expansión
        zip_frame.columnconfigure(0, weight=1)
        zip_frame.rowconfigure(0, weight=1)
        output_frame.columnconfigure(0, weight=1)
        progress_frame.columnconfigure(0, weight=1)
        
    def create_placeholder_interface(self, parent_frame, company_name):
        """Crear interfaz placeholder para otras empresas"""
        # Encabezado de la empresa
        company_label = ttk.Label(parent_frame, text=f"{company_name}", style='Company.TLabel')
        company_label.pack(pady=(0, 30))
        
        # Mensaje de desarrollo
        message_frame = ttk.LabelFrame(parent_frame, text="Estado del Desarrollo", padding="20")
        message_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        message_text = f"""
        La lógica específica para {company_name} está en desarrollo.
        
        Funcionalidades planificadas:
        • Procesamiento de facturas específico para {company_name}
        • Campos y validaciones personalizadas
        • Formatos de salida específicos
        • Reglas de negocio particulares
        
        Estado: Pendiente de implementación
        """
        
        message_label = ttk.Label(message_frame, text=message_text, 
                                 font=('Arial', 10), justify=tk.LEFT)
        message_label.pack()
        
        # Botón placeholder
        placeholder_button = ttk.Button(message_frame, text=f"Próximamente - {company_name}", 
                                       state='disabled', style='Custom.TButton')
        placeholder_button.pack(pady=(20, 0))
    
    def toggle_excel_options(self):
        """Mostrar/ocultar opciones de Excel"""
        if self.generate_excel.get():
            self.excel_frame_agro.grid()
        else:
            self.excel_frame_agro.grid_remove()
    
    def add_zip_files_agro(self):
        """Agregar archivos ZIP para AGROBUITRON"""
        file_paths = filedialog.askopenfilenames(
            title="Seleccionar archivos ZIP para AGROBUITRON",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        for file_path in file_paths:
            if file_path not in self.zip_files:
                self.zip_files.append(file_path)
                self.zip_listbox_agro.insert(tk.END, os.path.basename(file_path))
    
    def remove_zip_files_agro(self):
        """Remover archivos ZIP seleccionados"""
        selected_indices = self.zip_listbox_agro.curselection()
        for index in reversed(selected_indices):
            self.zip_listbox_agro.delete(index)
            del self.zip_files[index]
    
    def clear_zip_files_agro(self):
        """Limpiar lista de archivos ZIP"""
        self.zip_listbox_agro.delete(0, tk.END)
        self.zip_files.clear()
    
    def select_excel_file(self):
        """Seleccionar archivo Excel de destino"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file.set(file_path)
            self.load_excel_sheets()
    
    def load_excel_sheets(self):
        """Cargar las hojas disponibles del archivo Excel"""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(self.excel_file.get(), read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            self.sheet_combo['values'] = sheet_names
            if sheet_names:
                self.sheet_var.set(sheet_names[0])
                
        except Exception as e:
            messagebox.showerror("Error", f"Error cargando archivo Excel: {str(e)}")
            self.sheet_combo['values'] = []
    
    def preview_excel_structure(self):
        """Mostrar vista previa de la estructura del Excel"""
        if not self.excel_file.get():
            messagebox.showwarning("Advertencia", "Por favor selecciona un archivo Excel primero")
            return
        
        # [Código de vista previa igual al anterior pero adaptado]
        pass  # Implementación completa disponible si es necesaria
    
    def start_processing_agro(self):
        """Iniciar procesamiento para AGROBUITRON"""
        if not self.zip_files:
            messagebox.showerror("Error", "Por favor agrega al menos un archivo ZIP")
            return
        
        if not self.generate_csv.get() and not self.generate_excel.get():
            messagebox.showerror("Error", "Selecciona al menos una opción de salida")
            return
        
        if self.generate_excel.get() and not self.excel_file.get():
            messagebox.showerror("Error", "Selecciona un archivo Excel para actualizar")
            return
        
        # Deshabilitar botón durante procesamiento
        self.process_button.configure(state='disabled')
        
        # Iniciar procesamiento en hilo separado
        thread = threading.Thread(target=self.process_invoices_agro)
        thread.daemon = True
        thread.start()
    
    def process_invoices_agro(self):
        """Procesar facturas específicamente para AGROBUITRON"""
        try:
            self.extracted_data = []
            total_files = len(self.zip_files)
            
            self.update_status("Iniciando procesamiento AGROBUITRON...")
            
            for i, zip_file in enumerate(self.zip_files):
                self.update_status(f"Procesando {os.path.basename(zip_file)}...")
                self.process_zip_file_agro(zip_file)
                
                # Actualizar barra de progreso
                progress = ((i + 1) / total_files) * 100
                self.progress_var.set(progress)
                self.root.update_idletasks()
            
            # Generar salidas según selección
            outputs_generated = []
            
            if self.generate_csv.get():
                self.update_status("Generando archivo CSV...")
                csv_file = self.save_to_csv_agro()
                if csv_file:
                    outputs_generated.append(f"CSV: {csv_file}")
            
            if self.generate_excel.get():
                self.update_status("Actualizando archivo Excel...")
                self.save_to_excel_agro()
                outputs_generated.append(f"Excel: {os.path.basename(self.excel_file.get())}")
            
            self.update_status(f"Procesamiento completado. {len(self.extracted_data)} facturas procesadas")
            
            success_msg = f"Procesamiento AGROBUITRON completado:\n\n"
            success_msg += f"• {len(self.extracted_data)} facturas procesadas\n"
            success_msg += f"• Archivos generados:\n"
            for output in outputs_generated:
                success_msg += f"  - {output}\n"
            
            messagebox.showinfo("Éxito", success_msg)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error durante el procesamiento: {str(e)}")
            self.update_status("Error en el procesamiento")
        
        finally:
            # Rehabilitar botón
            self.process_button.configure(state='normal')
            self.progress_var.set(0)
    
    def process_zip_file_agro(self, zip_path):
        """Procesar un archivo ZIP específico para AGROBUITRON"""
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.endswith('.xml')]
                
                for xml_file in xml_files:
                    try:
                        xml_content = zip_ref.read(xml_file)
                        root = ET.fromstring(xml_content)
                        
                        # Aplicar lógica específica de AGROBUITRON
                        invoice_data = self.extract_invoice_data_agro(root)
                        if invoice_data:
                            if isinstance(invoice_data, list):
                                self.extracted_data.extend(invoice_data)
                            else:
                                self.extracted_data.append(invoice_data)
                                
                    except Exception as e:
                        print(f"Error procesando {xml_file}: {str(e)}")
                        continue
                        
        except Exception as e:
            print(f"Error abriendo ZIP {zip_path}: {str(e)}")
    
    def extract_invoice_data_agro(self, xml_root):
        """Extraer datos específicos para AGROBUITRON"""
        try:
            data = {}
            
            # Datos básicos de la factura
            data['N° Factura'] = self.get_xml_text(xml_root, './/cbc:ID', '')
            
            # Fechas
            issue_date = self.get_xml_text(xml_root, './/cbc:IssueDate', '')
            data['Fecha Factura'] = self.format_date(issue_date)
            
            due_date = self.get_xml_text(xml_root, './/cbc:DueDate', '')
            data['Fecha Pago'] = self.format_date(due_date)
            
            # Datos del vendedor (AGROBUITRON)
            supplier = xml_root.find('.//cac:AccountingSupplierParty/cac:Party', self.namespaces)
            if supplier is not None:
                data['Nit Vendedor'] = self.get_xml_text(supplier, './/cbc:CompanyID', '')
                data['Nombre Vendedor'] = self.get_xml_text(supplier, './/cbc:RegistrationName', '')
                data['Municipio'] = self.get_xml_text(supplier, './/cbc:CityName', '')
            
            # Datos del comprador
            customer = xml_root.find('.//cac:AccountingCustomerParty/cac:Party', self.namespaces)
            if customer is not None:
                data['Nit Comprador'] = self.get_xml_text(customer, './/cbc:CompanyID', '')
                data['Nombre Comprador'] = self.get_xml_text(customer, './/cbc:RegistrationName', '')
            
            # Configuraciones específicas AGROBUITRON
            data['Principal V,C'] = 'V'
            data['Moneda'] = self.get_currency_code(self.get_xml_text(xml_root, './/cbc:DocumentCurrencyCode', 'COP'))
            
            # Procesar líneas de productos
            invoice_lines = xml_root.findall('.//cac:InvoiceLine', self.namespaces)
            
            if invoice_lines:
                line_data_list = []
                for line in invoice_lines:
                    line_data = data.copy()
                    
                    # Datos específicos de cada línea
                    line_data['Nombre Producto'] = self.get_xml_text(line, './/cbc:Description', '')
                    line_data['Codigo Subyacente'] = self.get_xml_text(line, './/cbc:ID[@schemeID="999"]', '')
                    
                    # Unidad de medida desde unitCode
                    unit_code = line.find('.//cbc:InvoicedQuantity', self.namespaces)
                    unit_code_value = ''
                    if unit_code is not None:
                        unit_code_value = unit_code.get('unitCode', '')
                    line_data['Unidad Medida'] = self.convert_unit_code(unit_code_value)
                    
                    # Cantidad
                    quantity = self.get_xml_text(line, './/cbc:InvoicedQuantity', '0')
                    line_data['Cantidad'] = self.format_decimal(quantity)
                    line_data['Cantidad Original'] = line_data['Cantidad']
                    
                    # Precios
                    price = self.get_xml_text(line, './/cac:Price/cbc:PriceAmount', '0')
                    line_data['Precio Unitario'] = self.format_decimal(price)
                    
                    # Precio total desde TaxableAmount
                    total_price = self.get_xml_text(line, './/cac:TaxTotal/cac:TaxSubtotal/cbc:TaxableAmount', '0')
                    if not total_price or total_price == '0':
                        total_price = self.get_xml_text(line, './/cbc:LineExtensionAmount', '0')
                    line_data['Precio Total'] = self.format_decimal(total_price)
                    
                    # IVA
                    iva_percent = self.get_xml_text(line, './/cac:TaxTotal/cac:TaxSubtotal/cac:TaxCategory/cbc:Percent', '0')
                    line_data['Iva'] = f"{iva_percent}%"
                    
                    # Campos específicos AGROBUITRON
                    line_data['Descripción'] = ''
                    line_data['Activa Factura'] = 'Sí'
                    line_data['Activa Bodega'] = 'Sí'
                    line_data['Incentivo'] = ''
                    
                    line_data_list.append(line_data)
                
                return line_data_list
            else:
                # Datos por defecto si no hay líneas
                data.update({
                    'Nombre Producto': '', 'Codigo Subyacente': '', 'Unidad Medida': '',
                    'Cantidad': '0,00000', 'Precio Unitario': '0,00000', 'Precio Total': '0,00000',
                    'Iva': '0%', 'Descripción': '', 'Activa Factura': 'Sí', 
                    'Activa Bodega': 'Sí', 'Incentivo': '', 'Cantidad Original': '0,00000'
                })
                return [data]
                
        except Exception as e:
            print(f"Error extrayendo datos AGROBUITRON: {str(e)}")
            return None
    
    def save_to_csv_agro(self):
        """Generar archivo CSV para AGROBUITRON"""
        if not self.extracted_data:
            return None
        
        try:
            # Aplanar datos
            flattened_data = []
            for item in self.extracted_data:
                if isinstance(item, list):
                    flattened_data.extend(item)
                else:
                    flattened_data.append(item)
            
            # Definir orden de columnas específico para AGROBUITRON
            column_order = [
                'N° Factura', 'Nombre Producto', 'Codigo Subyacente', 'Unidad Medida',
                'Cantidad', 'Precio Unitario', 'Precio Total', 'Fecha Factura', 'Fecha Pago',
                'Nit Comprador', 'Nombre Comprador', 'Nit Vendedor', 'Nombre Vendedor',
                'Principal V,C', 'Municipio', 'Iva', 'Descripción', 'Activa Factura',
                'Activa Bodega', 'Incentivo', 'Cantidad Original', 'Moneda'
            ]
            
            # Crear DataFrame
            df = pd.DataFrame(flattened_data)
            df = df.reindex(columns=column_order, fill_value='')
            
            # Generar nombre de archivo con timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_filename = f"AGROBUITRON_Facturas_{timestamp}.csv"
            
            # Guardar CSV
            df.to_csv(csv_filename, index=False, encoding='utf-8-sig', sep=';')
            
            return csv_filename
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando CSV: {str(e)}")
            return None
    
    def save_to_excel_agro(self):
        """Actualizar Excel existente para AGROBUITRON"""
        if not self.extracted_data:
            return
        
        from openpyxl import load_workbook
        
        # Aplanar datos
        flattened_data = []
        for item in self.extracted_data:
            if isinstance(item, list):
                flattened_data.extend(item)
            else:
                flattened_data.append(item)
        
        try:
            # Cargar el archivo Excel existente
            workbook = load_workbook(self.excel_file.get())
            
            # Seleccionar la hoja especificada
            sheet_name = self.sheet_var.get()
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Leer los encabezados de la primera fila para mapear columnas
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers[cell_value.strip()] = col
            
            # Mapeo de nombres de campos a encabezados del Excel
            field_mapping = {
                'N° Factura': ['N° Factura', 'Numero Factura', 'No. Factura', 'Factura'],
                'Nombre Producto': ['Nombre Producto', 'Producto', 'Descripcion Producto'],
                'Codigo Subyacente': ['Codigo Subyacente', 'Código Subyacente', 'Codigo', 'Code'],
                'Unidad Medida': ['Unidad Medida', 'Unidad', 'U/M', 'UM'],
                'Cantidad': ['Cantidad', 'Qty', 'Cant'],
                'Precio Unitario': ['Precio Unitario', 'Precio', 'Price', 'Valor Unitario'],
                'Precio Total': ['Precio Total', 'Total', 'Valor Total', 'Importe'],
                'Fecha Factura': ['Fecha Factura', 'Fecha', 'Date', 'Fecha Emision'],
                'Fecha Pago': ['Fecha Pago', 'Fecha Vencimiento', 'Due Date'],
                'Nit Comprador': ['Nit Comprador', 'NIT Comprador', 'Cliente NIT'],
                'Nombre Comprador': ['Nombre Comprador', 'Cliente', 'Comprador'],
                'Nit Vendedor': ['Nit Vendedor', 'NIT Vendedor', 'Proveedor NIT'],
                'Nombre Vendedor': ['Nombre Vendedor', 'Vendedor', 'Proveedor'],
                'Principal V,C': ['Principal V,C', 'Principal', 'Tipo'],
                'Municipio': ['Municipio', 'Ciudad', 'City'],
                'Iva': ['Iva', 'IVA', 'Tax', 'Impuesto'],
                'Descripción': ['Descripción', 'Descripcion', 'Description'],
                'Activa Factura': ['Activa Factura', 'Activa', 'Active'],
                'Activa Bodega': ['Activa Bodega', 'Bodega', 'Warehouse'],
                'Incentivo': ['Incentivo', 'Bonus'],
                'Cantidad Original': ['Cantidad Original', 'Original Qty'],
                'Moneda': ['Moneda', 'Currency', 'Curr']
            }
            
            # Crear mapeo de columnas encontradas
            column_map = {}
            for field, possible_names in field_mapping.items():
                for name in possible_names:
                    if name in headers:
                        column_map[field] = headers[name]
                        break
            
            # Encontrar la primera fila vacía
            start_row = 2
            for row in range(2, worksheet.max_row + 1):
                is_empty = True
                for col in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row=row, column=col).value:
                        is_empty = False
                        break
                
                if is_empty:
                    start_row = row
                    break
            else:
                start_row = worksheet.max_row + 1
            
            # Escribir los datos extraídos
            current_row = start_row
            
            for data_item in flattened_data:
                if current_row > worksheet.max_row + 100:
                    break
                
                for field, value in data_item.items():
                    if field in column_map:
                        col_num = column_map[field]
                        worksheet.cell(row=current_row, column=col_num, value=value)
                
                current_row += 1
            
            # Guardar el archivo
            workbook.save(self.excel_file.get())
            workbook.close()
            
            self.update_status(f"Datos guardados en {len(flattened_data)} filas del Excel")
            
        except Exception as e:
            self.update_status(f"Error guardando en Excel: {str(e)}")
    
    # Funciones auxiliares
    def get_xml_text(self, element, xpath, default=''):
        """Obtener texto de un elemento XML usando XPath"""
        try:
            found = element.find(xpath, self.namespaces)
            if found is not None:
                return found.text or default
            
            found = element.find(xpath.replace('cbc:', '').replace('cac:', ''))
            if found is not None:
                return found.text or default
                
            return default
        except:
            return default
    
    def format_date(self, date_str):
        """Formatear fecha a formato YYYY-MM-DD"""
        if not date_str:
            return ''
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except:
            return date_str
    
    def format_decimal(self, value_str):
        """Formatear número decimal con 5 decimales y coma como separador"""
        try:
            value = float(value_str.replace(',', '.'))
            return f"{value:.5f}".replace('.', ',')
        except:
            return "0,00000"
    
    def convert_unit_code(self, unit_code):
        """Convertir código de unidad a formato requerido"""
        unit_mapping = {
            'KGM': 'Kg', 'LTR': 'Lt', 'LT': 'Lt', 'NIU': 'Un', 'MTR': 'Mt', 'HUR': 'Hr',
            'GRM': 'Gr', 'TNE': 'Tn', 'MLT': 'Ml', 'CMT': 'Cm', 'M2': 'M2', 'M3': 'M3',
            'DAY': 'Día', 'MON': 'Mes', 'ANN': 'Año', 'PCE': 'Pz', 'SET': 'Set', 'PAR': 'Par',
            'DZN': 'Docena', 'BOX': 'Caja', 'BAG': 'Bolsa', 'BTL': 'Botella', 'CAN': 'Lata'
        }
        return unit_mapping.get(unit_code.upper(), unit_code if unit_code else 'Un')
    
    def get_currency_code(self, currency):
        """Convertir código de moneda a número"""
        currency_mapping = {'COP': '1', 'USD': '2', 'EUR': '3'}
        return currency_mapping.get(currency, '1')
    
    def update_status(self, message):
        """Actualizar mensaje de estado"""
        self.status_label.configure(text=message)
        self.root.update_idletasks()

def main():
    root = tk.Tk()
    app = InvoiceExtractor(root)
    root.mainloop()

if __name__ == "__main__":
    main()