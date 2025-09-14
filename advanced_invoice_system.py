# -*- coding: utf-8 -*-
"""
Sistema de Facturas Electrónicas
Login (SQLite) → App principal (AGROBUITRON) → Reportes (solo admin)

- Usuarios por defecto:
    admin / admin123        (admin)
    operador / FacturasElectronicas2024  (operator)
"""

# ==========================
#  IMPORTS
# ==========================
import os
import sys
import csv
import zipfile
import hashlib
import sqlite3
import threading
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font

import pandas as pd

# Dependencias opcionales
try:
    import requests
    from packaging import version
    HAS_REQUESTS = True
except Exception:
    HAS_REQUESTS = False


# ==========================
#  CONSTANTES / CONFIG
# ==========================
DB_FILE = "facturas_users.db"


# ==========================
#  UTILIDADES GENERALES
# ==========================
def human_size(nbytes: int) -> str:
    if nbytes is None:
        return "N/A"
    try:
        if nbytes > 1024 * 1024:
            return f"{nbytes / (1024 * 1024):.1f} MB"
        if nbytes > 1024:
            return f"{nbytes / 1024:.1f} KB"
        return f"{nbytes} B"
    except Exception:
        return "N/A"


# ==========================
#  AUTENTICACIÓN
# ==========================
class AuthenticationWindow:
    """Ventana de login + gestión de contraseña"""

    def __init__(self, parent_callback):
        self.parent_callback = parent_callback
        self.root = tk.Tk()
        self.root.title("Sistema de Facturas Electrónicas - Acceso")
        self.root.geometry("450x550")
        self.root.configure(bg="#2c3e50")
        self.root.resizable(False, False)

        # Estado
        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.new_password_var = tk.StringVar()
        self.confirm_password_var = tk.StringVar()

        self._center()
        self._setup_database()
        self._build_ui()

    # ---------- Setup ----------
    def _center(self):
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (450 // 2)
        y = (self.root.winfo_screenheight() // 2) - (550 // 2)
        self.root.geometry(f"450x550+{x}+{y}")

    def _setup_database(self):
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                user_type TEXT NOT NULL DEFAULT 'operator',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP
            )
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                company TEXT NOT NULL,
                filename TEXT NOT NULL,
                records_processed INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                file_size INTEGER
            )
            """
        )

        admin_hash = hashlib.sha256("admin123".encode()).hexdigest()
        operator_hash = hashlib.sha256("FacturasElectronicas2024".encode()).hexdigest()

        cur.execute(
            "INSERT OR IGNORE INTO users (username, password_hash, user_type) VALUES (?, ?, ?)",
            ("admin", admin_hash, "admin"),
        )
        cur.execute(
            "INSERT OR IGNORE INTO users (username, password_hash, user_type) VALUES (?, ?, ?)",
            ("operador", operator_hash, "operator"),
        )

        conn.commit()
        conn.close()

    # ---------- UI ----------
    def _build_ui(self):
        main = tk.Frame(self.root, bg="#2c3e50")
        main.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        title_font = font.Font(family="Arial", size=24, weight="bold")
        tk.Label(
            main, text="SISTEMA DE FACTURAS", font=title_font, fg="#ecf0f1", bg="#2c3e50"
        ).pack(pady=(0, 10))
        tk.Label(
            main,
            text="Extracción de Facturas Electrónicas DIAN",
            font=("Arial", 12),
            fg="#bdc3c7",
            bg="#2c3e50",
        ).pack(pady=(0, 40))

        login = tk.Frame(main, bg="#34495e", relief=tk.RAISED, bd=2)
        login.pack(pady=20, padx=20, fill=tk.BOTH)

        tk.Label(
            login, text="Iniciar Sesión", font=("Arial", 16, "bold"), fg="#ecf0f1", bg="#34495e"
        ).pack(pady=(20, 30))

        tk.Label(login, text="Usuario:", font=("Arial", 11), fg="#ecf0f1", bg="#34495e").pack(
            anchor=tk.W, padx=30
        )
        user_entry = tk.Entry(login, textvariable=self.username_var, font=("Arial", 12), width=25, relief=tk.FLAT, bd=5)
        user_entry.pack(pady=(5, 15), padx=30, ipady=8)

        tk.Label(login, text="Contraseña:", font=("Arial", 11), fg="#ecf0f1", bg="#34495e").pack(
            anchor=tk.W, padx=30
        )
        
        login_pass_row = tk.Frame(login, bg="#34495e")
        login_pass_row.pack(pady=(5, 20), padx=30, fill=tk.X)

        self.login_pass_entry = tk.Entry(
            login_pass_row, textvariable=self.password_var, font=("Arial", 12),
            width=25, show="*", relief=tk.FLAT, bd=5
        )
        self.login_pass_entry.pack(side=tk.LEFT, ipady=8)

        
        self.login_show_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            login_pass_row,
            text="Mostrar",
            variable=self.login_show_var,
            command=lambda: self._toggle_entry_visibility(self.login_pass_entry, self.login_show_var),
            bg="#34495e", fg="#ecf0f1", selectcolor="#34495e", activebackground="#34495e",
        ).pack(side=tk.LEFT, padx=10)

        btns = tk.Frame(login, bg="#34495e")
        btns.pack(pady=(0, 20))

        tk.Button(
            btns,
            text="INICIAR SESIÓN",
            command=self.login,
            font=("Arial", 11, "bold"),
            bg="#27ae60",
            fg="white",
            width=15,
            height=2,
            relief=tk.FLAT,
            cursor="hand2",
        ).pack(pady=5)

        tk.Button(
            btns,
            text="Cambiar Contraseña",
            command=self._show_change_password,
            font=("Arial", 10),
            bg="#f39c12",
            fg="white",
            width=18,
            height=1,
            relief=tk.FLAT,
            cursor="hand2",
        ).pack(pady=5)

        self.change_pass_frame = tk.Frame(main, bg="#34495e", relief=tk.RAISED, bd=2)

        user_entry.bind("<Return>", lambda e: self.login_pass_entry.focus())
        self.login_pass_entry.bind("<Return>", lambda e: self.login())
        user_entry.focus()
    def _toggle_entry_visibility(self, entry: tk.Entry, var: tk.BooleanVar):
        """Muestra u oculta el texto de un Entry según var (True=mostrar)."""
        entry.config(show='' if var.get() else '*')

    def _toggle_entries_visibility(self, entries: list[tk.Entry], var: tk.BooleanVar):
        """Muestra u oculta el texto de varios Entry según var (True=mostrar)."""
        for e in entries:
            e.config(show='' if var.get() else '*')

    def _show_change_password(self):
        if not self.username_var.get().strip():
            messagebox.showwarning("Advertencia", "Ingrese su nombre de usuario primero")
            return

        for w in self.change_pass_frame.winfo_children():
            w.destroy()
        self.change_pass_frame.pack(pady=20, padx=20, fill=tk.BOTH)

        tk.Label(
            self.change_pass_frame,
            text="Cambiar Contraseña",
            font=("Arial", 16, "bold"),
            fg="#ecf0f1",
            bg="#34495e",
        ).pack(pady=(20, 30))

        tk.Label(self.change_pass_frame, text="Contraseña Actual:", font=("Arial", 11), fg="#ecf0f1", bg="#34495e").pack(
            anchor=tk.W, padx=30
        )
        tk.Entry(
            self.change_pass_frame, textvariable=self.password_var, font=("Arial", 12), width=25, show="*", relief=tk.FLAT, bd=5
        ).pack(pady=(5, 15), padx=30, ipady=8)

        tk.Label(self.change_pass_frame, text="Nueva Contraseña:", font=("Arial", 11), fg="#ecf0f1", bg="#34495e").pack(
            anchor=tk.W, padx=30
        )
        tk.Entry(
            self.change_pass_frame, textvariable=self.new_password_var, font=("Arial", 12), width=25, show="*", relief=tk.FLAT, bd=5
        ).pack(pady=(5, 15), padx=30, ipady=8)

        tk.Label(
            self.change_pass_frame, text="Confirmar Nueva Contraseña:", font=("Arial", 11), fg="#ecf0f1", bg="#34495e"
        ).pack(anchor=tk.W, padx=30)
        tk.Entry(
            self.change_pass_frame,
            textvariable=self.confirm_password_var,
            font=("Arial", 12),
            width=25,
            show="*",
            relief=tk.FLAT,
            bd=5,
        ).pack(pady=(5, 20), padx=30, ipady=8)

        btns = tk.Frame(self.change_pass_frame, bg="#34495e")
        btns.pack(pady=(0, 20))

        tk.Button(
            btns,
            text="CAMBIAR",
            command=self.change_password,
            font=("Arial", 11, "bold"),
            bg="#e74c3c",
            fg="white",
            width=12,
            height=2,
            relief=tk.FLAT,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btns,
            text="CANCELAR",
            command=self._hide_change_password,
            font=("Arial", 11),
            bg="#7f8c8d",
            fg="white",
            width=12,
            height=2,
            relief=tk.FLAT,
            cursor="hand2",
        ).pack(side=tk.LEFT, padx=5)

    def _hide_change_password(self):
        self.change_pass_frame.pack_forget()
        self.new_password_var.set("")
        self.confirm_password_var.set("")

    # ---------- Lógica ----------
    def _auth(self, username: str, password: str):
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        pwd_hash = hashlib.sha256(password.encode()).hexdigest()
        cur.execute(
            "SELECT username, user_type FROM users WHERE username = ? AND password_hash = ?",
            (username, pwd_hash),
        )
        user = cur.fetchone()
        if user:
            cur.execute("UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE username = ?", (username,))
            conn.commit()
        conn.close()
        return user

    def change_password(self):
        username = self.username_var.get().strip()
        current_pass = self.password_var.get()
        new_pass = self.new_password_var.get()
        confirm_pass = self.confirm_password_var.get()

        if not all([username, current_pass, new_pass, confirm_pass]):
            messagebox.showerror("Error", "Todos los campos son obligatorios")
            return
        if new_pass != confirm_pass:
            messagebox.showerror("Error", "Las contraseñas nuevas no coinciden")
            return
        if len(new_pass) < 6:
            messagebox.showerror("Error", "La nueva contraseña debe tener al menos 6 caracteres")
            return
        if not self._auth(username, current_pass):
            messagebox.showerror("Error", "Contraseña actual incorrecta")
            return

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        new_hash = hashlib.sha256(new_pass.encode()).hexdigest()
        cur.execute("UPDATE users SET password_hash = ? WHERE username = ?", (new_hash, username))
        conn.commit()
        conn.close()

        messagebox.showinfo("Éxito", "Contraseña cambiada exitosamente")
        self._hide_change_password()

    def login(self):
        username = self.username_var.get().strip()
        password = self.password_var.get()

        if not username or not password:
            messagebox.showerror("Error", "Ingrese usuario y contraseña")
            return

        user = self._auth(username, password)
        if user:
            self.root.destroy()
            self.parent_callback(user[0], user[1])  # username, user_type
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")
            self.password_var.set("")

    def run(self):
        self.root.mainloop()


# ==========================
#  REPORTES (ADMIN)
# ==========================
class ReportsWindow:
    """Ventana de reportes generados"""

    def __init__(self, parent_window):
        self.parent = parent_window
        self.window = tk.Toplevel(parent_window.root)
        self.window.title("Panel de Administración - Reportes")
        self.window.geometry("800x600")
        self.window.configure(bg="#f0f0f0")

        self._center()
        self._build_ui()
        self._load_reports()

    def _center(self):
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - (800 // 2)
        y = (self.window.winfo_screenheight() // 2) - (600 // 2)
        self.window.geometry(f"800x600+{x}+{y}")

    def _build_ui(self):
        tk.Label(self.window, text="Panel de Administración - Reportes Generados",
                 font=("Arial", 16, "bold"), bg="#f0f0f0").pack(pady=20)

        table_frame = tk.Frame(self.window, bg="#f0f0f0")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))

        cols = ("Usuario", "Empresa", "Archivo", "Registros", "Fecha", "Tamaño")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=15)

        for c in cols:
            self.tree.heading(c, text=c)

        self.tree.column("Usuario", width=100)
        self.tree.column("Empresa", width=120)
        self.tree.column("Archivo", width=220)
        self.tree.column("Registros", width=90)
        self.tree.column("Fecha", width=160)
        self.tree.column("Tamaño", width=90)

        vbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)
        hbar.pack(side=tk.BOTTOM, fill=tk.X)

        btns = tk.Frame(self.window, bg="#f0f0f0")
        btns.pack(pady=10)
        tk.Button(btns, text="Actualizar", command=self._load_reports,
                  font=("Arial", 11), bg="#3498db", fg="white", width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(btns, text="Exportar", command=self._export_reports,
                  font=("Arial", 11), bg="#27ae60", fg="white", width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(btns, text="Cerrar", command=self.window.destroy,
                  font=("Arial", 11), bg="#e74c3c", fg="white", width=12).pack(side=tk.LEFT, padx=5)

    def _load_reports(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT username, company, filename, records_processed, created_at, file_size
            FROM reports
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()
        conn.close()

        for (username, company, filename, records, created_at, file_size) in rows:
            self.tree.insert(
                "",
                tk.END,
                values=(username, company, filename, records, created_at, human_size(file_size)),
            )

    def _export_reports(self):
        try:
            conn = sqlite3.connect(DB_FILE)
            df = pd.read_sql_query(
                """
                SELECT username, company, filename, records_processed, created_at, file_size
                FROM reports
                ORDER BY created_at DESC
                """,
                conn,
            )
            conn.close()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = f"Reportes_Sistema_{timestamp}.csv"
            df.to_csv(out, index=False, encoding="utf-8-sig", sep=";")
            messagebox.showinfo("Éxito", f"Reportes exportados a: {out}")
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando reportes: {str(e)}")


# ==========================
#  APP PRINCIPAL
# ==========================
class InvoiceExtractor:
    """Aplicación principal: pestañas + procesamiento AGROBUITRON"""

    def __init__(self, username: str, user_type: str):
        self.username = username
        self.user_type = user_type

        self.root = tk.Tk()
        self.root.title(f"Sistema de Facturas Electrónicas - Usuario: {username}")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f0f0")

        # Estado
        self.current_version = "2.1.0"
        self.excel_file = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.zip_files = []
        self.extracted_data = []
        self.current_company = "AGROBUITRON"

        # Namespaces UBL/DIAN
        self.namespaces = {
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
            "ext": "urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2",
            "sts": "dian:gov:co:facturaelectronica:Structures-2-1",
        }

        self._setup_styles()
        self._build_ui()
        self._check_updates_optional()

    # ---------- UI ----------
    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Arial", 16, "bold"), background="#f0f0f0")
        style.configure("Subtitle.TLabel", font=("Arial", 10, "bold"), background="#f0f0f0")
        style.configure("Custom.TButton", font=("Arial", 10))
        style.configure("Company.TLabel", font=("Arial", 12, "bold"), foreground="#2E8B57")

    def _build_ui(self):
        # Header
        top = tk.Frame(self.root, bg="#34495e", height=60)
        top.pack(fill=tk.X)
        top.pack_propagate(False)

        tk.Label(
            top,
            text=f"Usuario: {self.username} | Tipo: {self.user_type.upper()}",
            font=("Arial", 12, "bold"),
            fg="white",
            bg="#34495e",
        ).pack(side=tk.LEFT, padx=20, pady=15)

        if self.user_type == "admin":
            tk.Button(
                top, text="Ver Reportes", command=self._open_reports, font=("Arial", 10), bg="#f39c12", fg="white"
            ).pack(side=tk.RIGHT, padx=10, pady=15)

        tk.Button(
            top, text="Cerrar Sesión", command=self._logout, font=("Arial", 10), bg="#e74c3c", fg="white"
        ).pack(side=tk.RIGHT, padx=10, pady=15)

        # Tabs
        main = ttk.Frame(self.root, padding="20")
        main.pack(fill=tk.BOTH, expand=True)  
        ttk.Label(main, text="Sistema de Extracción de Facturas Electrónicas", style="Title.TLabel").grid(
            row=0, column=0, columnspan=3, pady=(0, 20)
        )

        self.notebook = ttk.Notebook(main)
        self.notebook.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))

        self.agrobuitron_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.agrobuitron_frame, text="AGROBUITRON")

        self.mg_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.mg_frame, text="MG")

        self.rosas_frame = ttk.Frame(self.notebook, padding="15")
        self.notebook.add(self.rosas_frame, text="ROSAS")

        self._build_agrobuitron_tab()
        self._build_placeholder_tab(self.mg_frame, "MG")
        self._build_placeholder_tab(self.rosas_frame, "ROSAS")

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)

    def _open_reports(self):
        if self.user_type == "admin":
            ReportsWindow(self)
        else:
            messagebox.showerror("Error", "Acceso denegado")

    def _logout(self):
        self.root.destroy()
        main()  # Volver a login

    def _on_tab_changed(self, event):
        self.current_company = event.widget.tab("current")["text"]

    def _build_placeholder_tab(self, parent, company_name):
        ttk.Label(parent, text=f"{company_name}", style="Company.TLabel").pack(pady=(0, 30))
        lf = ttk.LabelFrame(parent, text="Estado del Desarrollo", padding="20")
        lf.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        msg = (
            f"La lógica específica para {company_name} está en desarrollo.\n\n"
            "Funcionalidades planificadas:\n"
            f"• Procesamiento específico para {company_name}\n"
            "• Validaciones personalizadas\n• Formatos de salida\n• Reglas de negocio\n\n"
            "Estado: Pendiente"
        )
        ttk.Label(lf, text=msg, font=("Arial", 10), justify=tk.LEFT).pack()
        ttk.Button(lf, text=f"Próximamente - {company_name}", state="disabled", style="Custom.TButton").pack(pady=20)

    def _build_agrobuitron_tab(self):
        frame = self.agrobuitron_frame
        ttk.Label(frame, text="AGROBUITRON SAS", style="Company.TLabel").grid(row=0, column=0, columnspan=3, pady=(0, 15))

        # Archivos ZIP
        zip_frame = ttk.LabelFrame(frame, text="Archivos ZIP con Facturas", padding="10")
        zip_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))

        self.zip_listbox_agro = tk.Listbox(zip_frame, height=8, selectmode=tk.MULTIPLE)
        self.zip_listbox_agro.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10), pady=10)

        vbar = ttk.Scrollbar(zip_frame, orient=tk.VERTICAL, command=self.zip_listbox_agro.yview)
        vbar.grid(row=0, column=2, sticky=(tk.N, tk.S), pady=10)
        self.zip_listbox_agro.configure(yscrollcommand=vbar.set)

        btns = ttk.Frame(zip_frame)
        btns.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E))
        ttk.Button(btns, text="Agregar ZIPs", command=self._add_zip_agro, style="Custom.TButton").grid(row=0, column=0, padx=5)
        ttk.Button(btns, text="Remover Seleccionados", command=self._remove_zip_agro, style="Custom.TButton").grid(row=0, column=1, padx=5)
        ttk.Button(btns, text="Limpiar Lista", command=self._clear_zip_agro, style="Custom.TButton").grid(row=0, column=2, padx=5)

        # Opciones salida
        out_frame = ttk.LabelFrame(frame, text="Opciones de Salida", padding="10")
        out_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))

        self.generate_csv = tk.BooleanVar(value=True)
        self.generate_excel = tk.BooleanVar(value=False)

        ttk.Checkbutton(out_frame, text="Generar archivo CSV", variable=self.generate_csv).grid(
            row=0, column=0, sticky=tk.W, padx=(0, 20)
        )
        ttk.Checkbutton(out_frame, text="Actualizar Excel existente", variable=self.generate_excel).grid(
            row=0, column=1, sticky=tk.W
        )

        self.excel_frame = ttk.Frame(out_frame)
        self.excel_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))

        ttk.Entry(self.excel_frame, textvariable=self.excel_file, width=50, state="readonly").grid(row=0, column=0, padx=(0, 10))
        ttk.Button(self.excel_frame, text="Seleccionar Excel", command=self._select_excel, style="Custom.TButton").grid(row=0, column=1, padx=5)
        ttk.Button(self.excel_frame, text="Vista Previa", command=self._preview_excel, style="Custom.TButton").grid(row=0, column=2, padx=5)

        sheet_frame = ttk.Frame(self.excel_frame)
        sheet_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        ttk.Label(sheet_frame, text="Hoja:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, width=20, state="readonly")
        self.sheet_combo.grid(row=0, column=1, padx=5, sticky=tk.W)

        self._toggle_excel_options()
        self.generate_excel.trace("w", lambda *_: self._toggle_excel_options())

        # Progreso
        prog = ttk.Frame(frame)
        prog.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        ttk.Label(prog, text="Progreso:", style="Subtitle.TLabel").grid(row=0, column=0, sticky=tk.W)
        self.progress_var = tk.DoubleVar()
        ttk.Progressbar(prog, variable=self.progress_var, maximum=100, length=400).grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        self.status_label = ttk.Label(prog, text="Listo para procesar")
        self.status_label.grid(row=2, column=0, sticky=tk.W, pady=(5, 0))

        # Procesar
        pbtn = ttk.Frame(frame)
        pbtn.grid(row=4, column=0, columnspan=3, pady=(0, 15))
        self.process_button = ttk.Button(pbtn, text="PROCESAR FACTURAS AGROBUITRON", command=self._start_process_agro, style="Custom.TButton")
        self.process_button.grid(row=0, column=0, padx=10)

        # Grid weights
        zip_frame.columnconfigure(0, weight=1)
        zip_frame.rowconfigure(0, weight=1)

    # ---------- Acciones UI ----------
    def _toggle_excel_options(self):
        if self.generate_excel.get():
            self.excel_frame.grid()
        else:
            self.excel_frame.grid_remove()

    def _add_zip_agro(self):
        files = filedialog.askopenfilenames(
            title="Seleccionar archivos ZIP para AGROBUITRON", filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        for f in files:
            if f not in self.zip_files:
                self.zip_files.append(f)
                self.zip_listbox_agro.insert(tk.END, os.path.basename(f))

    def _remove_zip_agro(self):
        sel = self.zip_listbox_agro.curselection()
        for idx in reversed(sel):
            self.zip_listbox_agro.delete(idx)
            del self.zip_files[idx]

    def _clear_zip_agro(self):
        self.zip_listbox_agro.delete(0, tk.END)
        self.zip_files.clear()

    def _select_excel(self):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.excel_file.set(path)
            self._load_excel_sheets()

    def _load_excel_sheets(self):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.excel_file.get(), read_only=True)
            names = wb.sheetnames
            wb.close()
            self.sheet_combo["values"] = names
            if names:
                self.sheet_var.set(names[0])
        except Exception as e:
            messagebox.showerror("Error", f"Error cargando archivo Excel: {str(e)}")
            self.sheet_combo["values"] = []

    def _preview_excel(self):
        if not self.excel_file.get():
            messagebox.showwarning("Advertencia", "Selecciona un archivo Excel primero")
            return
        messagebox.showinfo("Vista previa", f"Archivo seleccionado:\n{self.excel_file.get()}")

    # ---------- Proceso principal ----------
    def _start_process_agro(self):
        if not self.zip_files:
            messagebox.showerror("Error", "Agrega al menos un archivo ZIP")
            return
        if not self.generate_csv.get() and not self.generate_excel.get():
            messagebox.showerror("Error", "Selecciona al menos una opción de salida")
            return
        if self.generate_excel.get() and not self.excel_file.get():
            messagebox.showerror("Error", "Selecciona un archivo Excel para actualizar")
            return

        self.process_button.configure(state="disabled")
        t = threading.Thread(target=self._process_agro_thread, daemon=True)
        t.start()

    def _process_agro_thread(self):
        try:
            self.extracted_data = []
            total = len(self.zip_files)
            self._status("Iniciando procesamiento AGROBUITRON...")

            for i, z in enumerate(self.zip_files):
                self._status(f"Procesando {os.path.basename(z)}...")
                self._process_zip_agro(z)
                self.progress_var.set(((i + 1) / total) * 100)
                self.root.update_idletasks()

            outputs = []
            records = len(self.extracted_data)

            if self.generate_csv.get():
                self._status("Generando archivo CSV...")
                csv_path = self._save_csv_agro()
                if csv_path:
                    outputs.append(f"CSV: {csv_path}")
                    size = None
                    try:
                        size = os.path.getsize(csv_path)
                    except Exception:
                        pass
                    self._save_report(self.current_company, csv_path, records, size)

            if self.generate_excel.get():
                self._status("Actualizando Excel...")
                if self._save_excel_agro():
                    outputs.append(f"Excel: {os.path.basename(self.excel_file.get())}")
                    self._save_report(self.current_company, os.path.basename(self.excel_file.get()), records, None)

            self._status(f"Procesamiento completado. {records} facturas procesadas")
            msg = f"Procesamiento AGROBUITRON completado:\n\n• {records} facturas procesadas"
            if outputs:
                msg += "\n• Archivos generados:\n" + "\n".join([f"  - {o}" for o in outputs])
            messagebox.showinfo("Éxito", msg)

        except Exception as e:
            messagebox.showerror("Error", f"Error durante el procesamiento: {str(e)}")
            self._status("Error en el procesamiento")
        finally:
            self.process_button.configure(state="normal")
            self.progress_var.set(0)

    def _process_zip_agro(self, zip_path: str):
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                xmls = [f for f in zf.namelist() if f.lower().endswith(".xml")]
                for x in xmls:
                    try:
                        root = ET.fromstring(zf.read(x))
                        data = self._extract_agro(root)
                        if data:
                            if isinstance(data, list):
                                self.extracted_data.extend(data)
                            else:
                                self.extracted_data.append(data)
                    except Exception as e:
                        print(f"Error procesando {x}: {e}")
        except Exception as e:
            print(f"Error abriendo ZIP {zip_path}: {e}")

    def _extract_agro(self, xml_root):
        try:
            base = {}
            base["N° Factura"] = self._xml_text(xml_root, ".//cbc:ID", "")

            issue_date = self._xml_text(xml_root, ".//cbc:IssueDate", "")
            base["Fecha Factura"] = self._fmt_date(issue_date)

            due_date = self._xml_text(xml_root, ".//cbc:DueDate", "")
            base["Fecha Pago"] = self._fmt_date(due_date)

            supplier = xml_root.find(".//cac:AccountingSupplierParty/cac:Party", self.namespaces)
            if supplier is not None:
                base["Nit Vendedor"] = self._xml_text(supplier, ".//cbc:CompanyID", "")
                base["Nombre Vendedor"] = self._xml_text(supplier, ".//cbc:RegistrationName", "")
                base["Municipio"] = self._xml_text(supplier, ".//cbc:CityName", "")

            customer = xml_root.find(".//cac:AccountingCustomerParty/cac:Party", self.namespaces)
            if customer is not None:
                base["Nit Comprador"] = self._xml_text(customer, ".//cbc:CompanyID", "")
                base["Nombre Comprador"] = self._xml_text(customer, ".//cbc:RegistrationName", "")

            base["Principal V,C"] = "V"
            base["Moneda"] = self._currency_code(self._xml_text(xml_root, ".//cbc:DocumentCurrencyCode", "COP"))

            lines = xml_root.findall(".//cac:InvoiceLine", self.namespaces)
            if lines:
                out = []
                for ln in lines:
                    row = base.copy()
                    row["Nombre Producto"] = self._xml_text(ln, ".//cbc:Description", "")
                    row["Codigo Subyacente"] = self._xml_text(ln, './/cbc:ID[@schemeID="999"]', "")

                    unit = ln.find(".//cbc:InvoicedQuantity", self.namespaces)
                    unit_code = unit.get("unitCode", "") if unit is not None else ""
                    row["Unidad Medida"] = self._unit_code(unit_code)

                    qty = self._xml_text(ln, ".//cbc:InvoicedQuantity", "0")
                    row["Cantidad"] = self._fmt_dec(qty)
                    row["Cantidad Original"] = row["Cantidad"]

                    price = self._xml_text(ln, ".//cac:Price/cbc:PriceAmount", "0")
                    row["Precio Unitario"] = self._fmt_dec(price)

                    total = self._xml_text(ln, ".//cac:TaxTotal/cac:TaxSubtotal/cbc:TaxableAmount", "0")
                    if not total or total == "0":
                        total = self._xml_text(ln, ".//cbc:LineExtensionAmount", "0")
                    row["Precio Total"] = self._fmt_dec(total)

                    iva = self._xml_text(ln, ".//cac:TaxTotal/cac:TaxSubtotal/cac:TaxCategory/cbc:Percent", "0")
                    row["Iva"] = f"{iva}%"

                    row["Descripción"] = ""
                    row["Activa Factura"] = "Sí"
                    row["Activa Bodega"] = "Sí"
                    row["Incentivo"] = ""

                    out.append(row)
                return out
            else:
                base.update(
                    {
                        "Nombre Producto": "",
                        "Codigo Subyacente": "",
                        "Unidad Medida": "",
                        "Cantidad": "0,00000",
                        "Precio Unitario": "0,00000",
                        "Precio Total": "0,00000",
                        "Iva": "0%",
                        "Descripción": "",
                        "Activa Factura": "Sí",
                        "Activa Bodega": "Sí",
                        "Incentivo": "",
                        "Cantidad Original": "0,00000",
                    }
                )
                return [base]
        except Exception as e:
            print(f"Error extrayendo datos AGROBUITRON: {e}")
            return None

    # ---------- Salidas ----------
    def _save_csv_agro(self):
        if not self.extracted_data:
            return None
        try:
            rows = []
            for x in self.extracted_data:
                if isinstance(x, list):
                    rows.extend(x)
                else:
                    rows.append(x)

            cols = [
                "N° Factura",
                "Nombre Producto",
                "Codigo Subyacente",
                "Unidad Medida",
                "Cantidad",
                "Precio Unitario",
                "Precio Total",
                "Fecha Factura",
                "Fecha Pago",
                "Nit Comprador",
                "Nombre Comprador",
                "Nit Vendedor",
                "Nombre Vendedor",
                "Principal V,C",
                "Municipio",
                "Iva",
                "Descripción",
                "Activa Factura",
                "Activa Bodega",
                "Incentivo",
                "Cantidad Original",
                "Moneda",
            ]

            df = pd.DataFrame(rows)
            df = df.reindex(columns=cols, fill_value="")

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = f"AGROBUITRON_Facturas_{ts}.csv"
            df.to_csv(out, index=False, encoding="utf-8-sig", sep=";")
            return out
        except Exception as e:
            messagebox.showerror("Error", f"Error generando CSV: {str(e)}")
            return None

    def _save_excel_agro(self) -> bool:
        if not self.extracted_data:
            return False
        try:
            from openpyxl import load_workbook

            rows = []
            for x in self.extracted_data:
                rows.extend(x if isinstance(x, list) else [x])

            wb = load_workbook(self.excel_file.get())
            ws = wb[self.sheet_var.get()] if (self.sheet_var.get() and self.sheet_var.get() in wb.sheetnames) else wb.active

            # Encabezados → posición de columna
            headers = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=c).value
                if val:
                    headers[str(val).strip()] = c

            field_mapping = {
                "N° Factura": ["N° Factura", "Numero Factura", "No. Factura", "Factura"],
                "Nombre Producto": ["Nombre Producto", "Producto", "Descripcion Producto"],
                "Codigo Subyacente": ["Codigo Subyacente", "Código Subyacente", "Codigo", "Code"],
                "Unidad Medida": ["Unidad Medida", "Unidad", "U/M", "UM"],
                "Cantidad": ["Cantidad", "Qty", "Cant"],
                "Precio Unitario": ["Precio Unitario", "Precio", "Price", "Valor Unitario"],
                "Precio Total": ["Precio Total", "Total", "Valor Total", "Importe"],
                "Fecha Factura": ["Fecha Factura", "Fecha", "Date", "Fecha Emision"],
                "Fecha Pago": ["Fecha Pago", "Fecha Vencimiento", "Due Date"],
                "Nit Comprador": ["Nit Comprador", "NIT Comprador", "Cliente NIT"],
                "Nombre Comprador": ["Nombre Comprador", "Cliente", "Comprador"],
                "Nit Vendedor": ["Nit Vendedor", "NIT Vendedor", "Proveedor NIT"],
                "Nombre Vendedor": ["Nombre Vendedor", "Vendedor", "Proveedor"],
                "Principal V,C": ["Principal V,C", "Principal", "Tipo"],
                "Municipio": ["Municipio", "Ciudad", "City"],
                "Iva": ["Iva", "IVA", "Tax", "Impuesto"],
                "Descripción": ["Descripción", "Descripcion", "Description"],
                "Activa Factura": ["Activa Factura", "Activa", "Active"],
                "Activa Bodega": ["Activa Bodega", "Bodega", "Warehouse"],
                "Incentivo": ["Incentivo", "Bonus"],
                "Cantidad Original": ["Cantidad Original", "Original Qty"],
                "Moneda": ["Moneda", "Currency", "Curr"],
            }

            colmap = {}
            for field, names in field_mapping.items():
                for n in names:
                    if n in headers:
                        colmap[field] = headers[n]
                        break

            # Primera fila vacía
            start = None
            for r in range(2, ws.max_row + 2):
                if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
                    start = r
                    break
            if start is None:
                start = ws.max_row + 1

            rcur = start
            for item in rows:
                for field, val in item.items():
                    if field in colmap:
                        ws.cell(row=rcur, column=colmap[field], value=val)
                rcur += 1

            wb.save(self.excel_file.get())
            wb.close()
            self._status(f"Datos guardados en {len(rows)} filas del Excel")
            return True
        except Exception as e:
            self._status(f"Error guardando en Excel: {str(e)}")
            return False

    # ---------- Persistencia de Reporte ----------
    def _save_report(self, company: str, filename: str, records: int, file_size: int | None):
        try:
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO reports (username, company, filename, records_processed, file_size) VALUES (?, ?, ?, ?, ?)",
                (self.username, company, filename, records, file_size),
            )
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Error guardando reporte: {e}")

    # ---------- Helpers XML / formatos ----------
    def _xml_text(self, el, xpath: str, default=""):
        try:
            found = el.find(xpath, self.namespaces)
            if found is not None and found.text is not None:
                return found.text
            simple = xpath.replace("cbc:", "").replace("cac:", "").replace("ext:", "").replace("sts:", "")
            found = el.find(simple)
            if found is not None and found.text is not None:
                return found.text
            return default
        except Exception:
            return default

    def _fmt_date(self, s: str) -> str:
        if not s:
            return ""
        try:
            datetime.strptime(s, "%Y-%m-%d")
            return s
        except Exception:
            return s

    def _fmt_dec(self, s: str) -> str:
        try:
            val = float(str(s).replace(",", "."))
            return f"{val:.5f}".replace(".", ",")
        except Exception:
            return "0,00000"

    def _unit_code(self, code: str) -> str:
        mapping = {
            "KGM": "Kg",
            "LTR": "Lt",
            "LT": "Lt",
            "NIU": "Un",
            "MTR": "Mt",
            "HUR": "Hr",
            "GRM": "Gr",
            "TNE": "Tn",
            "MLT": "Ml",
            "CMT": "Cm",
            "M2": "M2",
            "M3": "M3",
            "DAY": "Día",
            "MON": "Mes",
            "ANN": "Año",
            "PCE": "Pz",
            "SET": "Set",
            "PAR": "Par",
            "DZN": "Docena",
            "BOX": "Caja",
            "BAG": "Bolsa",
            "BTL": "Botella",
            "CAN": "Lata",
        }
        return mapping.get((code or "").upper(), code if code else "Un")

    def _currency_code(self, cur: str) -> str:
        return {"COP": "1", "USD": "2", "EUR": "3"}.get(cur, "1")

    def _status(self, msg: str):
        self.status_label.configure(text=msg)
        self.root.update_idletasks()

    # ---------- Actualizaciones (opcional) ----------
    def _check_updates_optional(self):
        if not HAS_REQUESTS:
            return
        try:
            url = "https://api.github.com/repos/LuisVeraVR/operator-auto/releases/latest"
            r = requests.get(url, timeout=5)
            if r.status_code == 200:
                rel = r.json()
                latest = (rel.get("tag_name") or "").replace("v", "")
                if latest and version.parse(latest) > version.parse(self.current_version):
                    if messagebox.askyesno("Actualización Disponible", f"Nueva versión: {latest}\n¿Desea descargarla?"):
                        assets = rel.get("assets") or []
                        if assets:
                            self._download_update(assets[0].get("browser_download_url") or "")
                        else:
                            messagebox.showinfo(
                                "Actualización", "No hay binarios adjuntos. Visite el repositorio para descargar."
                            )
        except Exception:
            pass  # silencioso

    def _download_update(self, url: str):
        if not (HAS_REQUESTS and url):
            messagebox.showinfo(
                "Actualización", "Descargue manualmente desde:\nhttps://github.com/LuisVeraVR/operator-auto/releases"
            )
            return
        try:
            resp = requests.get(url, stream=True)
            out = "invoice_extractor_update.exe"
            with open(out, "wb") as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)
            messagebox.showinfo("Actualización", f"Descargado como {out}.\nCierre la app y ejecute el nuevo archivo.")
        except Exception as e:
            messagebox.showerror("Error", f"Error descargando actualización: {str(e)}")


# ==========================
#  ENTRYPOINT
# ==========================
def main():
    def on_login_ok(username: str, user_type: str):
        app = InvoiceExtractor(username, user_type)
        app.root.mainloop()

    auth = AuthenticationWindow(on_login_ok)
    auth.run()


if __name__ == "__main__":
    main()
